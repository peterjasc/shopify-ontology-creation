import pandas as pd
import openpyxl
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import LabelEncoder
import joblib
import os
import sys
import numpy as np
from sentence_transformers import SentenceTransformer
import torch

TFIDF_VECTORIZER = None
LR_MODEL = None
LABEL_ENCODER_INFERENCE = None
CONFIDENCE_THRESHOLD = 0.20

class Colors:
 RESET = "\033[0m"
 RED = "\033[31m"
 GREEN = "\033[32m"
 YELLOW = "\033[33m"
 BLUE = "\033[34m"
 MAGENTA = "\033[35m"
 CYAN = "\033[36m"
 WHITE = "\033[37m"
 BOLD = "\033[1m"

MODEL_DIR = "./bow_collection_model"

def load_bow_model_for_inference():
    global TFIDF_VECTORIZER, LR_MODEL, LABEL_ENCODER_INFERENCE
    if not os.path.exists(MODEL_DIR):
        print(f"Error: Trained BoW model not found at '{MODEL_DIR}'. Please train the model first.", file=sys.stderr)
        return False
    try:
        TFIDF_VECTORIZER = joblib.load(os.path.join(MODEL_DIR, "tfidf_vectorizer.joblib"))
        LABEL_ENCODER_INFERENCE = joblib.load(os.path.join(MODEL_DIR, "label_encoder.joblib"))
        LR_MODEL = joblib.load(os.path.join(MODEL_DIR, "logistic_regression_model.joblib"))
        print(f"Successfully loaded BoW model from {MODEL_DIR}")
        return True
    except Exception as e:
        print(f"Error loading BoW model: {e}. Ensure the model was saved correctly.", file=sys.stderr)
        return False

def assign_collection_with_bow(title):
    if not title or TFIDF_VECTORIZER is None or LR_MODEL is None or LABEL_ENCODER_INFERENCE is None:
        return "Unassigned"
    try:
        title_vec = TFIDF_VECTORIZER.transform([str(title)])
        probabilities = LR_MODEL.predict_proba(title_vec)[0]
        max_confidence = np.max(probabilities)
        print(max_confidence)
        if max_confidence < CONFIDENCE_THRESHOLD:
            return "Miscellaneous"
        else:
            predicted_label_id = np.argmax(probabilities)
            predicted_collection = LABEL_ENCODER_INFERENCE.inverse_transform([predicted_label_id])[0]
            return predicted_collection
    except Exception as e:
        print(f"Warning: BoW prediction failed for title '{title}': {e}. Assigning 'Unassigned'.", file=sys.stderr)
        return "Unassigned"

def add_color(color, text):
 print(f"{color}{text}{Colors.RESET}", end="")

def open_file(file_path):
 abs_path = os.path.abspath(file_path)
 print(f"{Colors.BLUE}Attempting to open file: {abs_path}{Colors.RESET}")
 try:
  if sys.platform.startswith('darwin'):
   import subprocess
   subprocess.Popen(['open', abs_path])
  elif sys.platform.startswith('win'):
   os.startfile(abs_path)
  elif sys.platform.startswith('linux'):
   import subprocess
   subprocess.Popen(['xdg-open', abs_path])
  else:
   raise OSError(f"Unsupported operating system: {sys.platform}")
  print(f"{Colors.GREEN}Successfully attempted to open: {abs_path}{Colors.RESET}")
  return None
 except Exception as e:
  return f"Failed to open file {abs_path}: {e}"

def get_excel_column_number(sheet, search_value):
 for col_idx, cell in enumerate(sheet[1]):
  if str(cell.value).strip().lower() == str(search_value).strip().lower():
   return col_idx + 1, True, None
 return 0, False, f"Column with header '{search_value}' not found in the first row."

def process_pdfs_in_xlsx(input_filepath, output_filepath,
       row_to_start_from, title_column_header,
       collection_column_header):
 workbook = None
 try:
  workbook = openpyxl.load_workbook(input_filepath)
  sheet = workbook.active
  print(f"{Colors.BLUE}Processing sheet: {sheet.title}{Colors.RESET}")
  title_col_idx, found_title_col, err_title_col = get_excel_column_number(sheet, title_column_header)
  if not found_title_col:
   return f"{Colors.RED}Error: {err_title_col}{Colors.RESET}"
  print(f"{Colors.GREEN}Value '{title_column_header}' found in column number: {title_col_idx}{Colors.RESET}")
  collection_col_idx, found_collection_col, err_collection_col = get_excel_column_number(sheet,
                          collection_column_header)
  if not found_collection_col:
   return f"{Colors.RED}Error: {err_collection_col}{Colors.RESET}"
  print(
   f"{Colors.GREEN}Value '{collection_column_header}' found in column number: {collection_col_idx}{Colors.RESET}")
  for row_index in range(row_to_start_from, sheet.max_row + 1):
   title_value = sheet.cell(row=row_index, column=title_col_idx).value
   title_desc_value = sheet.cell(row=row_index, column=3).value
   assigned_collection = assign_collection_with_bow(title_value)
   sheet.cell(row=row_index, column=collection_col_idx, value=assigned_collection)
   print(
    f"{Colors.CYAN}Row {row_index}: Title='{title_desc_value}' -> BOW Assigned Collection='{assigned_collection}'{Colors.RESET}")
  workbook.save(output_filepath)
  print(f"\n{Colors.GREEN}Processing complete. Modified file saved to: {output_filepath}{Colors.RESET}")
  return None
 except FileNotFoundError:
  return f"{Colors.RED}Error: Input file not found at '{input_filepath}'. Please check the path.{Colors.RESET}"
 except Exception as e:
  return f"{Colors.RED}A general error occurred during XLSX processing: {e}{Colors.RESET}"
 finally:
  if workbook:
   pass

def main():
 if not load_bow_model_for_inference():
  print(
   f"{Colors.RED}BERT model not loaded. Please train the model using the 'Conceptual BERT Training Script' first, and ensure it's saved to '{MODEL_DIR}'.{Colors.RESET}")
  sys.exit(1)
 default_xlsx_path = "Tooted.xlsx"
 default_output_xlsx_path = "TootedOutput.xlsx"
 print(f"Operating System: {sys.platform}")
 print(f"Default path for input XLSX file: {default_xlsx_path}")
 add_color(Colors.BLUE, "Enter the path to the input XLSX file (press Enter for default): ")
 input_xlsx_file = input().strip()
 final_input_xlsx_filepath = input_xlsx_file if input_xlsx_file else default_xlsx_path
 print(f"Using input path: {final_input_xlsx_filepath}")
 print(f"Default path for output XLSX file: {default_output_xlsx_path}")
 add_color(Colors.BLUE, "Enter the path for the output XLSX file (press Enter for default): ")
 output_xlsx_file = input().strip()
 final_output_xlsx_filepath = output_xlsx_file if output_xlsx_file else default_output_xlsx_path
 print(f"Using output path: {final_output_xlsx_filepath}")
 title_column_header = "Long PD"
 collection_column_header = "Collection"
 default_row_to_start_from = 2
 add_color(Colors.BLUE,
     f"Enter the row number to start processing from (1-based, press Enter for default: {default_row_to_start_from}): ")
 row_input = input().strip()
 row_to_start_from = default_row_to_start_from
 if row_input:
  try:
   parsed_row = int(row_input)
   if parsed_row < 1:
    print(
     f"{Colors.YELLOW}Row number must be 1 or greater. Using default: {default_row_to_start_from}{Colors.RESET}")
   else:
    row_to_start_from = parsed_row
  except ValueError:
   print(
    f"{Colors.YELLOW}Invalid input '{row_input}' for row number. Using default: {default_row_to_start_from}{Colors.RESET}")
 print(f"Processing Starts From Row (1-based): {row_to_start_from}")
 print(f"\nFinal Input File: {final_input_xlsx_filepath}")
 print(f"Final Output File: {final_output_xlsx_filepath}")
 print(f"Title Column Header: '{title_column_header}'")
 print(f"Collection Column Header: '{collection_column_header}'")
 print(f"Processing Starts From Row (1-based): {row_to_start_from}")
 print(f"{Colors.BOLD}{Colors.BLUE}Starting XLSX and PDF processing...{Colors.RESET}")
 err_msg = process_pdfs_in_xlsx(
  final_input_xlsx_filepath,
  final_output_xlsx_filepath,
  row_to_start_from,
  title_column_header,
  collection_column_header
 )
 if err_msg:
  print(f"\n{Colors.BOLD}{Colors.RED}Application terminated with error: {err_msg}{Colors.RESET}")
  sys.exit(1)
 print(f"\n{Colors.BOLD}{Colors.GREEN}All operations completed successfully!{Colors.RESET}")

if __name__ == "__main__":
 main()